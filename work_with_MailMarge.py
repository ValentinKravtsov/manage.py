from openpyxl import load_workbook
import warnings


def bd_retail_stations():
    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
    wb = load_workbook('./Mail Merge.xlsx')
    sheet = wb['RBA']
    bd = {}

    r = 2  # строка в экселе
    while True:
        retail = sheet.cell(row=r, column=10).value
        work = sheet.cell(row=r, column=26).value
        r += 1
        if retail == None:
            break
        elif (work.lower().strip() == 'no') or (retail in bd):
            continue
        else:
            stations = []

            r2 = 2  # строка в экселе
            while True:
                s = r2
                compare_retail = sheet.cell(row=r2, column=10).value
                station = sheet.cell(row=s, column=2).value
                work = sheet.cell(row=r2, column=26).value
                r2 += 1
                if compare_retail == None:
                    break
                if (work.lower().strip() == 'no') or (compare_retail != retail):
                    continue
                stations.append(station)
            bd[retail] = stations

    return bd

# rt_st = bd_retail_stations()
# count = 0
# for i in rt_st:
#     stations = rt_st[i]
#     count += len(stations)
# print(count)

    # print(i)
    # print(rt_st[i])
    # for station in stations:
    #     print(station)